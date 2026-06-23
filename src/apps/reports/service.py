from datetime import date
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from sqlalchemy.orm import selectinload
from jinja2 import Environment, FileSystemLoader
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import tempfile

from src.apps.finance.models import Invoice, InvoiceLineItem
from src.apps.procurement.models import PurchaseOrder, POItem, Vendor
from src.apps.boq.models import BOQItem, BudgetVersion
from src.apps.projects.models import Project
from src.apps.tenancy.models import Tenant
from src.core.exceptions import NotFoundError
from src.core.config import settings

TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "templates")
jinja_env = Environment(loader=FileSystemLoader(TEMPLATE_DIR))


class ReportService:
    def __init__(self, db: AsyncSession, tenant: Tenant):
        self.db = db
        self.tenant = tenant

    def _scope(self, model):
        return and_(model.tenant_id == self.tenant.id, model.deleted_at.is_(None))

    def _render(self, template_name: str, **context) -> str:
        template = jinja_env.get_template(template_name)
        return template.render(
            primary_color=self.tenant.primary_color or "#2563eb",
            generated_at=date.today().isoformat(),
            tenant=self.tenant,
            **context,
        )

    def _html_to_pdf(self, html: str) -> bytes:
        return HTML(string=html).write_pdf()

    # ── Invoice PDF ─────────────────────────────────────────────

    async def invoice_pdf(self, invoice_id: str) -> bytes:
        result = await self.db.execute(
            select(Invoice)
            .options(selectinload(Invoice.line_items))
            .where(and_(Invoice.id == invoice_id, self._scope(Invoice)))
        )
        invoice = result.scalar_one_or_none()
        if not invoice:
            raise NotFoundError("Invoice")

        vendor = None
        if invoice.vendor_id:
            v_result = await self.db.execute(
                select(Vendor).where(and_(
                    Vendor.id == invoice.vendor_id, self._scope(Vendor)
                ))
            )
            vendor = v_result.scalar_one_or_none()

        html = self._render("invoice.html", invoice=invoice, vendor=vendor)
        return self._html_to_pdf(html)

    # ── Purchase Order PDF ──────────────────────────────────────

    async def po_pdf(self, po_id: str) -> bytes:
        result = await self.db.execute(
            select(PurchaseOrder)
            .options(selectinload(PurchaseOrder.items))
            .where(and_(PurchaseOrder.id == po_id, self._scope(PurchaseOrder)))
        )
        po = result.scalar_one_or_none()
        if not po:
            raise NotFoundError("Purchase Order")

        v_result = await self.db.execute(
            select(Vendor).where(and_(
                Vendor.id == po.vendor_id, self._scope(Vendor)
            ))
        )
        vendor = v_result.scalar_one_or_none()
        if not vendor:
            raise NotFoundError("Vendor")

        html = self._render("purchase_order.html", po=po, vendor=vendor)
        return self._html_to_pdf(html)

    # ── BOQ PDF ─────────────────────────────────────────────────

    async def boq_pdf(self, version_id: str, project_id: str) -> bytes:
        v_result = await self.db.execute(
            select(BudgetVersion).where(and_(
                BudgetVersion.id == version_id,
                BudgetVersion.project_id == project_id,
                self._scope(BudgetVersion),
            ))
        )
        version = v_result.scalar_one_or_none()
        if not version:
            raise NotFoundError("Budget Version")

        items_result = await self.db.execute(
            select(BOQItem).where(and_(
                BOQItem.budget_version_id == version_id,
                BOQItem.project_id == project_id,
                self._scope(BOQItem),
            )).order_by(BOQItem.sort_order)
        )
        items = items_result.scalars().all()

        p_result = await self.db.execute(
            select(Project).where(and_(
                Project.id == project_id, self._scope(Project)
            ))
        )
        project = p_result.scalar_one_or_none()
        if not project:
            raise NotFoundError("Project")

        html = self._render("boq.html", version=version, items=items, project=project)
        return self._html_to_pdf(html)

    # ── BOQ Excel ───────────────────────────────────────────────

    async def boq_excel(self, version_id: str, project_id: str) -> bytes:
        v_result = await self.db.execute(
            select(BudgetVersion).where(and_(
                BudgetVersion.id == version_id,
                BudgetVersion.project_id == project_id,
                self._scope(BudgetVersion),
            ))
        )
        version = v_result.scalar_one_or_none()
        if not version:
            raise NotFoundError("Budget Version")

        items_result = await self.db.execute(
            select(BOQItem).where(and_(
                BOQItem.budget_version_id == version_id,
                BOQItem.project_id == project_id,
                self._scope(BOQItem),
            )).order_by(BOQItem.sort_order)
        )
        items = items_result.scalars().all()

        wb = Workbook()
        ws = wb.active
        ws.title = "BOQ"

        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Title
        ws.merge_cells("A1:I1")
        ws["A1"] = f"Bill of Quantities — {version.name} v{version.version_number}"
        ws["A1"].font = Font(bold=True, size=14, color="2563EB")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:I2")
        ws["A2"] = f"Grand Total: {version.currency} {version.grand_total:,.2f}"
        ws["A2"].font = Font(size=10, color="6B7280")
        ws["A2"].alignment = Alignment(horizontal="center")

        headers = [
            "Item #", "Description", "Unit", "Quantity", "Unit Rate",
            "Amount", "Material", "Labour", "Equipment",
        ]
        col_widths = [10, 45, 8, 12, 12, 14, 12, 12, 12]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        row = 5
        for item in items:
            values = [
                item.item_number,
                item.description,
                item.unit,
                item.quantity,
                item.unit_rate,
                item.amount,
                item.material_rate,
                item.labour_rate,
                item.equipment_rate,
            ]
            is_section = item.is_section_header
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border
                if is_section:
                    cell.font = Font(bold=True, size=10)
                    cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
                if col_idx >= 4 and not is_section:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0.00"
            row += 1

        # Totals row
        row += 1
        total_labels = [
            "", "Total", "", "",
            sum(i.quantity * i.unit_rate for i in items if not i.is_section_header),
            sum(i.amount for i in items if not i.is_section_header),
            sum(i.material_rate for i in items if not i.is_section_header),
            sum(i.labour_rate for i in items if not i.is_section_header),
            sum(i.equipment_rate for i in items if not i.is_section_header),
        ]
        for col_idx, val in enumerate(total_labels, 1):
            cell = ws.cell(row=row, column=col_idx, value=val if val else "")
            cell.font = Font(bold=True, size=10)
            cell.border = thin_border
            if isinstance(val, (int, float)) and val:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0.00"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name

        with open(tmp_path, "rb") as f:
            data = f.read()
        os.unlink(tmp_path)
        return data
